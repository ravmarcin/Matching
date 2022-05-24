import pandas as pd
from googletrans import Translator
import nltk.corpus
import nltk.tokenize
import nltk.stem.snowball
from nltk.corpus import wordnet
import nltk
import string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook


def import_xlsx(in_dir):
    """
    Function to import the data from the xlsx file
    :arg in_dir: string     - directory of the input xlsx file
    :return df: dataframe   - pandas DF
    """
    try:
        if isinstance(in_dir, str):
            df = pd.read_excel(io=in_dir, sheet_name=0)
            return df
        else:
            print('Wrong format of the given parameter')
    except KeyError:
        print('No such file in directory or the file is in another format')


def save_xlsx(df, in_dir='', out_dir=''):
    """
    Function to save the data into the xlsx file
    :arg df: dataframe      - pandas DF
    :arg in_dir: string     - directory of the input xlsx file
    :arg out_dir: string    - directory of the output xlsx file
    :return -
    """
    try:
        if isinstance(out_dir, str):
            wb = load_workbook(in_dir)  # load as openpyxl workbook; useful to keep the original layout
            # which is discarded in the following dataframe
            ws = wb.active
            rows = dataframe_to_rows(df, index=False)
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            wb.save(out_dir)
        else:
            print('Wrong format of the given parameter')
    except KeyError:
        print('No such directory or the ')


def check_if_new_rec(new_ver_dir='', temp_dir='', col_check=1, update=False):
    """
    Function to check if there are any new records
    :arg new_ver_dir: string            - directory of the new xlsx file
    :arg temp_dir: string               - directory of the temporary xlsx file
    :arg col_check: int                 - the column number for filter
    :arg update: boolean                - if True, then temporary xlsx file will be update
    :return new_check_df: dataframe     - new records given as pandas df
    """
    old_ver = import_xlsx(in_dir=temp_dir)
    old_ver_cols = old_ver.columns.values.tolist()
    old_ver_check = old_ver[old_ver_cols[col_check]].values.tolist()
    new_ver = import_xlsx(in_dir=new_ver_dir)
    new_ver_cols = new_ver.columns.values.tolist()
    new_ver_check = new_ver[new_ver_cols[col_check]].values.tolist()
    new_check_ls = []
    for i in range(len(new_ver_check)):
        if new_ver_check[i] not in old_ver_check:
            new_check_ls.append(new_ver_check[i])
    new_check_df = new_ver[new_ver[new_ver_cols[col_check]].isin(new_check_ls)]
    if update:
        save_xlsx(df=new_ver, out_dir=temp_dir)
    return new_check_df


def check_if_assign(df_dir='', col_check=14, no_stat='No'):
    """
    Function to check if there are records without assign
    :arg df_dir: string         - directory of the xlsx file
    :arg col_check: int         - the column number for assign check
    :arg no_stat: string        - status for 'No' assign
    :return df_check dataframe  - filtered records given as pandas df
    """
    df = import_xlsx(in_dir=df_dir)
    df_fill = df.fillna(no_stat)
    df_cols = df_fill.columns.values.tolist()
    df_check = df_fill.loc[df_fill[df_cols[col_check]] == no_stat]
    return df_check


def get_mentee(df, val, col=0):
    """
    Function to get mentee dataframe based on the index
    :arg df: dataframe          - input dataframe
    :arg val: str               - value to look for
    :arg col: int               - the column where to look for
    :return df_n: dataframe     - reduced pandas df
    """
    df_cols = df.columns.values.tolist()
    df_n = pd.DataFrame(columns=df_cols)
    df_n = pd.concat([df_n, df.loc[df[df_cols[col]] == val]], ignore_index=False)
    return df_n


def reduce_df(df, params=None, id_col=1):
    """
    Function to reduce the dataframe for only needed columns
    :arg df: dataframe          - input dataframe
    :arg id_col: int            - the column for id
    :arg params: list of int    - list of columns to export
    :return df_n: dataframe     - reduced columns given as pandas df
    """
    df_cols = df.columns.values.tolist()
    df_n = pd.DataFrame(df[df_cols[id_col]])
    if params is None:
        # [language, new role, theme, location]
        params = [5, 9, 10, 12]
    if isinstance(params, list):
        for i in range(len(params)):
            df_n.insert(i + 1, df_cols[params[i]], df[df_cols[params[i]]].values.tolist())
        return df_n


def trans_df(df, lang='en', cols=None, trans_col_names=False):
    """
    Function to translate the dataframe
    :arg df: dataframe              - input dataframe
    :arg lang: string               - a language to which you want to translate
    :arg cols: list of int          - list of columns to translate
    :arg trans_col_names: boolean   - reduced columns given as pandas df
    :return df: dataframe           - translated pandas df
    """
    if cols is None:
        cols = [1, 2, 3]
    col_ls = df.columns.values.tolist()
    if trans_col_names:
        translator = Translator()
        col_trans_ls = [translator.translate(txt, dest=lang).text for txt in col_ls]
        df.set_axis(col_trans_ls, axis=1, inplace=True)
    for col in cols:
        df = trans_col(df=df, col=col, lang=lang)
    return df


def trans_col(df, col=0, lang='en'):
    """
    Function to translate a column in the dataframe
    :arg df: dataframe              - input dataframe
    :arg col: int                   - a column number to translate
    :arg lang: string               - a language to which you want to translate
    :return df: dataframe           - pandas df with translated column
    """
    col_ls = df.columns.values.tolist()
    sen_ls = df[col_ls[col]].values.tolist()
    sen_trans_ls = []
    for i in range(len(sen_ls)):
        txt = str(sen_ls[i])
        translator = Translator()
        if len(txt) < 800:
            txt_trans = translator.translate(txt, dest=lang).text
        sen_trans_ls.append(txt_trans)
    df[col_ls[col]] = df[col_ls[col]].replace(sen_ls, sen_trans_ls)
    return df


def filter_by(df, f_dir, col=1):
    """
    Function to filter the dataframe based on the column and configuration file with filter values
    :arg df: dataframe              - input dataframe
    :arg col: int                   - a column number for filter
    :arg f_dir: string              - a directory for configuration file with filter values
    :return words_uq: list          - a list of unique values from the filtered column in relation to the file
    :return words_ls: list          - a list of lists of  values from the filtered column in relation to the file
    :return df: dataframe           - pandas df with filtered column
    """
    df_n = df
    f = open(f_dir, 'r')
    file = f.readlines()
    f.close()
    file = list(map(lambda x: x.lower()[:-1], file))
    col_ls = df_n.columns.values.tolist()
    word_ls, words_uq, words_ls = [], [], []
    for txt in df_n[col_ls[col]]:
        txt_sep = tokenize(str(txt))
        txt_sep = list(map(lambda x: x.lower(), txt_sep))
        word_n = []
        for word in txt_sep:
            if word in file:
                word_n.append(word)
                if word not in words_uq:
                    words_uq.append(word)
        if len(word_n) == 0:
            word_n = ['Nan']
            words_ls.append(word_n)
        else:
            word_n = [" ".join(word_n)]
            words_ls.append(word_n)
        word_ls.append(word_n)
    if len(word_ls) > 1:
        df_n[col_ls[col]] = df_n[col_ls[col]].replace(df_n[col_ls[col]].values.tolist(), word_ls)
    else:
        df_n[col_ls[col]] = df_n[col_ls[col]].replace(df_n[col_ls[col]].values.tolist(), word_ls[0])
    return words_uq, words_ls, df_n


def strong_match(df_to_fil, df_from_to, f_dir, words_uq, words_ls, by_col_to=1, col_from=0, col_to=0):
    """
    Function to match based on the exact words (strong matching)
    :arg df_to_fil: dataframe                       - input dataframe to match
    :arg df_from_to: dataframe                      - input dataframe from where the values will be get for matching
    :arg f_dir: string                              - a directory for configuration file with filter values
    :arg words_uq: list of string                   - a list of unique values from the column in relation to the file
    :arg words_ls: list of string                   - a list of lists of values from the column in relation to the file
    :arg by_col_to: int                             - column number for which the matching will be made
    :arg col_to: int                                - column number for <to> id values
    :arg col_from: int                              - column number for <from> id values
    :return df_from_to_red_ls: list of dataframes   - list of matched <from> values per <to> value in pandas df format
    """
    f = open(f_dir, 'r')
    file = f.readlines()
    f.close()
    file = list(map(lambda x: x.lower()[:-1], file))

    col_from_to_ls = df_from_to.columns.values.tolist()
    col_to_ls = df_from_to[col_from_to_ls[col_to]].values.tolist()
    col_to_fil_ls = df_to_fil.columns.values.tolist()
    uq_to_fil_ls = []
    from_to_ls_match = []
    from_to_ls_match_ls = []
    for txt in df_to_fil[col_to_fil_ls[by_col_to]]:
        txt_sep = tokenize(str(txt))
        txt_sep = list(map(lambda x: x.lower(), txt_sep))
        uq_to_fil = []
        from_to_ls = []
        from_to_ls_ls = []
        for word in txt_sep:
            if word in file:
                uq_to_fil.append(word)
                if word in words_uq:
                    for i in range(len(words_ls)):
                        if word in words_ls[i][0].split():
                            from_to_ls.append(col_to_ls[i])
        if len(uq_to_fil) == 0:
            uq_to_fil = ['Nan']
        else:
            uq_to_fil = [" ".join(uq_to_fil)]
        if len(from_to_ls) == 0:
            from_to_ls = 'Nan'
            from_to_ls_ls.append([from_to_ls])
        else:
            from_to_ls_ls.append(from_to_ls)
            from_to_ls = ','.join(from_to_ls)
        uq_to_fil_ls.append(uq_to_fil)
        from_to_ls_match.append(from_to_ls)
        from_to_ls_match_ls.append(from_to_ls_ls)
    df_cols = df_from_to.columns.values.tolist()
    df_from_to_red_ls = []
    for i in range(len(from_to_ls_match_ls)):
        df_i = pd.DataFrame(columns=df_cols)
        for j in range(len(from_to_ls_match_ls[i])):
            for k in range(len(from_to_ls_match_ls[i][j])):
                df_i = pd.concat([df_i, df_from_to.loc[df_from_to[df_cols[col_from]] == from_to_ls_match_ls[i][j][k]]],
                                 ignore_index=True)
        df_from_to_red_ls.append(df_i)
    return df_from_to_red_ls


def get_stopwords():
    """
    Function to get default English stopwords and extend with punctuation
    :arg
    :return stopwords: object - ntlk stopwords object
    """
    stopwords = nltk.corpus.stopwords.words('english')
    stopwords.extend(string.punctuation)
    stopwords.append('')
    return stopwords


def get_lemmae():
    """
    Function to get default Lemmatizer
    :arg
    :return lemmatizer: object - ntlk Lemmatizer object
    """
    lemmatizer = nltk.stem.wordnet.WordNetLemmatizer()
    return lemmatizer


def tokenize(text):
    """
    Function to tokenize the given text
    :arg text: str                  - sentence
    :return tokens: list of str     - tokenized list of words
    """
    tokens = nltk.word_tokenize(text)
    return tokens


def get_wordnet_pos(pos_tag):
    """
    Function to get a word type in wordnet format of the given word
    :arg pos_tag: tuple of str          - word + word type
    :return pos_tag_wn: tuple of str    - word + word type in wordnet format
    """
    if pos_tag[1].startswith('J'):
        pos_tag_wn = (pos_tag[0], wordnet.ADJ)
    elif pos_tag[1].startswith('V'):
        pos_tag_wn = (pos_tag[0], wordnet.VERB)
    elif pos_tag[1].startswith('N'):
        pos_tag_wn = (pos_tag[0], wordnet.NOUN)
    elif pos_tag[1].startswith('R'):
        pos_tag_wn = (pos_tag[0], wordnet.ADV)
    else:
        pos_tag_wn = (pos_tag[0], wordnet.NOUN)
    return pos_tag_wn


def is_ci_lemma_stopwords_set_match(a, b):
    """
    Function to compute the Jaccard similarity between two sentences
    :arg a: str                 - sentence a
    :arg b: str                 - sentence b
    :return ratio: float        - Jaccard similarity ratio
    """
    lemmatizer = get_lemmae()
    stopwords = get_stopwords()
    pos_a = map(get_wordnet_pos, nltk.pos_tag(tokenize(a)))
    pos_b = map(get_wordnet_pos, nltk.pos_tag(tokenize(b)))
    lemmae_a = [lemmatizer.lemmatize(token.lower().strip(string.punctuation), pos) for token, pos in pos_a
                if pos == wordnet.NOUN and token.lower().strip(string.punctuation) not in stopwords]
    lemmae_b = [lemmatizer.lemmatize(token.lower().strip(string.punctuation), pos) for token, pos in pos_b
                if pos == wordnet.NOUN and token.lower().strip(string.punctuation) not in stopwords]
    ratio = len(set(lemmae_a).intersection(lemmae_b)) / float(len(set(lemmae_a).union(lemmae_b)))
    return ratio


def light_match(df_to_fil, df_from_to_ls, by_cols_to=None, by_cols_from=None,
                prefix_='Ratio for: '):
    """
    Function to match based on the exact words (strong matching)
    :arg df_to_fil: dataframe                       - input dataframe to match
    :arg df_from_to_ls: list of dataframes          - input dataframes from where the values will be get for matching
    :arg by_cols_to: list of int                    - list of column numbers for <to>
                                                      for which the matching will be made
    :arg by_cols_from: list of int                  - list of column numbers for <from>
                                                      for which the matching will be made
    :arg prefix_: str                               - a prefix which will be added to a new columns
    :return df_from_to_red_ls: list of dataframes   - list of matched <from> values per <to> value in pandas df format
    """
    if by_cols_to is None or by_cols_from is None:
        by_cols_to = [3]
        by_cols_from = [3]
    if len(by_cols_to) == len(by_cols_from) and len(by_cols_to) > 0:
        col_to_ls = df_to_fil.columns.values.tolist()
        col_from_ls = df_from_to_ls[0].columns.values.tolist()
        df_from_to_n_ls = []
        to_idx = df_to_fil.index
        for i in range(len(to_idx)):
            df_from_to = df_from_to_ls[i]
            from_idx = df_from_to.index
            df_from_to_n = df_from_to.copy()
            row_sen_to = df_to_fil.loc[to_idx == to_idx[i]]
            for by_col in range(len(by_cols_to)):
                sen_to = row_sen_to[col_to_ls[by_cols_to[by_col]]].values.tolist()
                ratios = []
                for j in range(len(from_idx)):
                    row_sen_from = df_from_to.loc[from_idx == from_idx[j]]
                    sen_from = row_sen_from[col_from_ls[by_cols_from[by_col]]].values.tolist()
                    ratio = is_ci_lemma_stopwords_set_match(a=sen_to[0], b=sen_from[0])
                    ratios.append(ratio)
                new_col = prefix_ + col_to_ls[by_cols_to[by_col]] + ' & ' + col_from_ls[by_cols_from[by_col]]
                df_from_to_n[new_col] = ratios
            df_from_to_n_ls.append(df_from_to_n)
        return df_from_to_n_ls


def update_df(df, val_idx, val_to_fill, user='-', col_idx=1, col_to_fill=15, col_status=14, new_status='Yes',
              col_user=13):
    """
    Function to update the dataframes based on the new match
    :arg df: dataframe              - input dataframe to fill
    :arg val_idx: str               - value to look for
    :arg val_to_fill: str           - value to fill
    :arg user: str                  - user name
    :arg col_idx: int               - column where to look for value
    :arg col_to_fill: int           - column where to fill the new value
    :arg col_status: int            - column for status change
    :arg new_status: str            - new status
    :arg col_user: int              - col for filling the user name
    :return df_update: dataframe    - updated dataframe
    """
    df_update = df
    col_ls = df.columns.values.tolist()
    df_update.loc[df_update[col_ls[col_idx]] == val_idx, [col_ls[col_to_fill]]] = val_to_fill
    df_update.loc[df_update[col_ls[col_idx]] == val_idx, [col_ls[col_status]]] = new_status
    df_update.loc[df_update[col_ls[col_idx]] == val_idx, [col_ls[col_user]]] = user
    return df_update


def update_mentor_df(df, val_idx, val_to_fill, user='-', col_idx=1, col_to_fill=17,
                     col_user=14):
    """
    Function to update the dataframes based on the new match
    :arg df: dataframe              - input dataframe to fill
    :arg val_idx: str               - value to look for
    :arg val_to_fill: str           - value to fill
    :arg user: str                  - user name
    :arg col_idx: int               - column where to look for value
    :arg col_to_fill: int           - column where to fill the new value
    :arg col_status: int            - column for status change
    :arg new_status: str            - new status
    :arg col_user: int              - col for filling the user name
    :return df_update: dataframe    - updated dataframe
    """
    df_update = df
    col_ls = df.columns.values.tolist()
    val_is = str(df_update.loc[df_update[col_ls[col_idx]] == val_idx, [col_ls[col_to_fill]]].values.tolist()[0])
    df_update.loc[df_update[col_ls[col_idx]] == val_idx, [col_ls[col_to_fill]]] = val_is + ' ' + val_to_fill
    df_update.loc[df_update[col_ls[col_idx]] == val_idx, [col_ls[col_user]]] = user
    return df_update


def get_mentee_info(df, val, col=1):
    col_ls = df.columns.values.tolist()
    sel_row = df.loc[df[col_ls[col]] == val].copy()
    cols_to_trans = [5, 6, 7, 8, 9, 10, 11, 12]
    sel_row_trans = trans_df(sel_row, lang='en', cols=cols_to_trans, trans_col_names=False)
    output = {'name': sel_row_trans[col_ls[2]].values.tolist()[0],
              'role': sel_row_trans[col_ls[7]].values.tolist()[0],
              'linkedin': sel_row_trans[col_ls[3]].values.tolist()[0],
              'location': sel_row_trans[col_ls[12]].values.tolist()[0],
              'languages': sel_row_trans[col_ls[5]].values.tolist()[0],
              'experience': sel_row_trans[col_ls[8]].values.tolist()[0],
              'topic': sel_row_trans[col_ls[10]].values.tolist()[0]
              }
    return output


def get_mentor_info(df, val, col=1):
    col_ls = df.columns.values.tolist()
    sel_row = df.loc[df[col_ls[col]] == val].copy()
    cols_to_trans = [4, 8, 9, 10, 13]
    sel_row_trans = trans_df(sel_row, lang='en', cols=cols_to_trans, trans_col_names=False)
    output = {'name': sel_row_trans[col_ls[2]].values.tolist()[0],
              'role': sel_row_trans[col_ls[8]].values.tolist()[0],
              'linkedin': sel_row_trans[col_ls[5]].values.tolist()[0],
              'location': sel_row_trans[col_ls[13]].values.tolist()[0],
              'languages': sel_row_trans[col_ls[4]].values.tolist()[0],
              'experience': sel_row_trans[col_ls[9]].values.tolist()[0],
              'topic': sel_row_trans[col_ls[10]].values.tolist()[0]
              }
    return output


def check_if_busy(df, idx_col=1, col_max=11, col_add=17):
    col_ls = df.columns.values.tolist()
    df_n = pd.DataFrame(columns=col_ls)
    idx_ls = df[col_ls[idx_col]].values.tolist()
    max_ls = df[col_ls[col_max]].values.tolist()
    add_ls = df[col_ls[col_add]].values.tolist()
    for i in range(len(max_ls)):
        max_i = 10
        try:
            max_i = int(max_ls[i])
        except:
            pass
        txt = str(add_ls[i])
        add_i = 0
        if len(txt) > 0:
            add_i = len(txt.split())
        if add_i < max_i:
            df_n = pd.concat([df_n, df.loc[df[col_ls[idx_col]] == idx_ls[i]]], ignore_index=False)
    return df_n


def check_dupl(df, idx_col=0):
    col_ls = df.columns.values.tolist()
    df_n = pd.DataFrame(columns=col_ls)
    idx_ls = df[col_ls[idx_col]].values.tolist()
    idx_ls_n = []
    df_index = df.index
    for i in range(len(idx_ls)):
        if idx_ls[i] not in idx_ls_n:
            df_n = pd.concat([df_n, df.loc[df.index == df_index[i]]], ignore_index=False)
            idx_ls_n.append(idx_ls[i])
    return df_n